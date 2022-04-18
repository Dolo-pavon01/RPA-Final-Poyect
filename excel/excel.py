
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill


class Excel:

    def __init__(self, file, max_col, data_only=True):

        self.__file = file
        self.__wb = load_workbook(file, data_only=data_only)

        self.__ws = self.__wb.active
        self.__header = None
        self.__style_header = None
        self.__max_col = max_col

    def rows(self):

        header_load = True

        for row in self.__ws.iter_rows(max_col=self.__max_col):

            row_dict = dict()

            if (header_load):
                self.__header = [cel.value for cel in row]
                cel = row[0]
                self.__style_header = [cel.font.color.rgb, cel.font.size,
                                       cel.font.bold, cel.fill.start_color, cel.fill.fill_type]
                header_load = False

            else:
                for cel, key in zip(row, self.__header):
                    row_dict[key] = cel.value
                yield row_dict

    def load_fil(self, list_dictionaries, keys):

        row_number = 2
        for dictionary in list_dictionaries:
            col_number = 1
            for key in keys:
                char = get_column_letter(col_number)
                self.__ws[char + str(row_number)] = dictionary[key]
                col_number += 1
            row_number += 1

    def write_header(self, list_titles):
        rgb, size, bold, start_color, type = self.__style_header
        row_number = 1
        col_number = 1
        for title in list_titles:
            char = get_column_letter(col_number)
            self.__ws[char + str(row_number)] = title
            if self.__style_header:
                cel = self.__ws[char + str(row_number)]
                cel.font = Font(size=size, color=rgb, bold=bold)
                cel.fill = PatternFill(
                    start_color=start_color, end_color=start_color, fill_type=type)
            col_number += 1
        self.__header = list_titles
        self.__max_col = len(list_titles)

    def add_cell(self, xrow, ycol, value):
        char = get_column_letter(ycol)
        self.__ws[char + str(xrow)] = value

    def header(self):
        return self.__header

    def save(self):
        self.__wb.save(self.__file)
