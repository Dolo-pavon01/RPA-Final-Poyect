
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


class Excel:

    def __init__(self, file,max_col, data_only =True):

        self.__file = file
        self.__wb = load_workbook(file,data_only= data_only)
        
        self.__ws = self.__wb.active
        self.__header = None
        self.__max_col = max_col

    def rows(self):

        header_load = True
       
        for row in self.__ws.iter_rows(max_col=self.__max_col):
            
            row_dict= dict()

            if (header_load): 
                self.__header = [cel.value for cel in row]
                header_load = False

            else:
                for cel,key in zip(row,self.__header):
                    row_dict[key] =cel.value
                yield row_dict
    
    def delete_row(self,num_row):

        self.__ws.delete_rows(num_row)

    def load_fil(self,list_dictionaries,keys):


        row_number = 2
        for dictionary in list_dictionaries:
            col_number = 1
            for key in keys:
                char = get_column_letter(col_number)
                self.__ws[char + str(row_number)] = dictionary[key]
                col_number+=1
            row_number+=1
  
    def header(self):
        return self.__header
        
    def save(self):
        self.__wb.save(self.__file)






    


